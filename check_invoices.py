#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script kiểm tra và lọc hóa đơn:
1. Món chưa được chuyển sang format "Tiếng Việt / Tiếng Anh"
2. Hóa đơn có bia/rượu đã được thay thế (bằng cách so sánh với menu gốc)
"""

import openpyxl
from pathlib import Path
import json
import re
import sys

# Import parse_menu để load menu
PACKAGE_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = PACKAGE_DIR
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from Menu.parse_menu import parse_excel_menu

def normalize_menu_key(s):
    """Chuẩn hóa chuỗi để so sánh tên món"""
    if not s:
        return ''
    s = s.lower().strip()
    # Bỏ ký tự không phải chữ/số/khoảng trắng
    s = re.sub(r'[^\w\s]', ' ', s)
    # Gom nhiều khoảng trắng liên tiếp thành 1
    s = re.sub(r'\s+', ' ', s)
    return s

def is_beverage_with_10_percent_tax(item_name):
    """Kiểm tra xem món có phải là bia, rượu, hoặc Coke 10% đường (không phải Coke Zero/Light) không
    
    Lưu ý: Coke thường có 10% đường nên tính thuế 10% (giống bia/rượu)
    Coke Light và Coke Zero có lượng đường < 10g nên tính thuế 8%
    """
    if not item_name:
        return False
    
    item_lower = item_name.lower()
    
    # Danh sách từ khóa cho bia/rượu
    beverage_keywords = ['beer', 'bia', 'wine', 'rượu', 'whiskey', 'vodka', 'rum', 'gin', 'tequila', 
                        'sake', 'soju', 'champagne', 'cocktail', 'martini', 'margarita']
    
    # Kiểm tra bia/rượu
    if any(kw in item_lower for kw in beverage_keywords):
        return True
    
    # Kiểm tra Coke (Coca-Cola) THƯỜNG - có 10% đường nên tính thuế 10% (giống bia/rượu)
    # Loại trừ: Coke Zero, Coke Light, và các biến thể ít đường/không đường
    if 'coke' in item_lower or 'coca' in item_lower:
        exclude_keywords = ['zero', 'light', 'ít đường', 'không đường', 'it duong', 'khong duong', 'less sugar', 'no sugar']
        is_coke_light_or_zero = any(exclude_kw in item_lower for exclude_kw in exclude_keywords)
        if not is_coke_light_or_zero:
            return True  # Coke thường (10% đường) tính thuế 10%
    
    return False

def load_menu_items():
    """Load tất cả menu items với giá gốc"""
    menu_dir = PROJECT_ROOT / "Menu"
    menu_files = [
        menu_dir / "simple-place-menu.xlsx",
        menu_dir / "taco-place-menu.xlsx"
    ]
    
    all_menu_items = []
    menu_by_price = {}  # price -> list of items (để tìm món bia/rượu theo giá)
    
    for menu_file in menu_files:
        if menu_file.exists():
            try:
                items = parse_excel_menu(str(menu_file))
                for item in items:
                    all_menu_items.append(item)
                    price = item.get('price', 0)
                    if price > 0:
                        if price not in menu_by_price:
                            menu_by_price[price] = []
                        menu_by_price[price].append(item)
            except Exception as e:
                print(f"⚠️  Lỗi khi load menu {menu_file.name}: {e}")
    
    return all_menu_items, menu_by_price

def find_original_beverage_in_menu(original_price, menu_by_price):
    """Tìm món bia/rượu/Coke trong menu có giá gốc
    
    Tìm kiếm với độ lệch cho phép:
    - Ưu tiên tìm chính xác
    - Nếu không có, tìm gần đúng (sai số <= 5000) để bao quát các trường hợp giá khác nhau
    """
    # Tìm chính xác giá
    if original_price in menu_by_price:
        for item in menu_by_price[original_price]:
            item_name = item.get('name', '')
            if is_beverage_with_10_percent_tax(item_name):
                return item
    
    # Tìm gần đúng (sai số <= 5000) - mở rộng để bao quát các giá khác nhau
    # Sắp xếp theo độ lệch nhỏ nhất trước
    candidates = []
    for price, items in menu_by_price.items():
        if abs(price - original_price) <= 5000:
            for item in items:
                item_name = item.get('name', '')
                if is_beverage_with_10_percent_tax(item_name):
                    diff = abs(price - original_price)
                    candidates.append((diff, item))
    
    # Trả về món có độ lệch nhỏ nhất
    if candidates:
        candidates.sort(key=lambda x: x[0])
        return candidates[0][1]
    
    return None

def check_invoices():
    """Kiểm tra tất cả hóa đơn trong tax_files bằng cách so sánh với menu gốc"""
    base_dir = Path(__file__).parent
    tax_dir = base_dir / "tax_files"
    
    if not tax_dir.exists():
        print("❌ Không tìm thấy thư mục tax_files")
        return {
            'invoices_without_format': [],
            'invoices_with_beverages': []
        }
    
    # Load menu
    print("📚 Đang load menu...")
    all_menu_items, menu_by_price = load_menu_items()
    print(f"   ✓ Đã load {len(all_menu_items)} món từ menu")
    
    # Tạo mapping tên món -> giá gốc từ menu
    menu_name_to_price = {}
    for item in all_menu_items:
        name = item.get('name', '')
        price = item.get('price', 0)
        if name and price > 0:
            # Map cả tên đầy đủ và phần tiếng Anh
            menu_name_to_price[normalize_menu_key(name)] = price
            if ' / ' in name:
                parts = name.split(' / ', 1)
                if len(parts) == 2:
                    english_part = parts[1].strip()
                    menu_name_to_price[normalize_menu_key(english_part)] = price
    
    # Kết quả
    invoices_without_format = []  # Hóa đơn có món chưa format
    invoices_with_beverages = []  # Hóa đơn có bia/rượu đã thay thế
    
    # Kiểm tra từng file hóa đơn
    print(f"\n🔍 Đang kiểm tra các file hóa đơn...")
    for invoice_file in sorted(tax_dir.glob("*.xlsx")):
        try:
            wb = openpyxl.load_workbook(invoice_file)
            ws = wb.active
            
            # Lấy số hóa đơn từ tên file (format: SỐ_HÓA_ĐƠN - ...)
            invoice_number = invoice_file.stem.split(' - ')[0] if ' - ' in invoice_file.stem else invoice_file.stem
            
            has_unformatted = False
            beverages_replaced = []
            
            # Kiểm tra từng món trong hóa đơn
            for row in range(2, ws.max_row + 1):
                product_name = ws.cell(row=row, column=3).value
                invoice_price = ws.cell(row=row, column=6).value
                
                if not product_name:
                    continue
                
                product_name_str = str(product_name)
                
                # Kiểm tra 1: Món chưa có format "Tiếng Việt / Tiếng Anh"
                if ' / ' not in product_name_str:
                    has_unformatted = True
                
                # Kiểm tra 2: Món có thể là bia/rượu đã được thay thế
                # So sánh giá với menu để phát hiện món thay thế
                if invoice_price and isinstance(invoice_price, (int, float)):
                    price_float = float(invoice_price)
                    price_int = int(price_float)
                    
                    # Bỏ qua nếu là phí dịch vụ
                    if 'phí dịch vụ' in product_name_str.lower() or 'service fee' in product_name_str.lower():
                        continue
                    
                    # Bỏ qua nếu món này là bia/rượu/Coke (không phải món thay thế)
                    if is_beverage_with_10_percent_tax(product_name_str):
                        continue
                    
                    # Tìm giá gốc của món này trong menu
                    product_key = normalize_menu_key(product_name_str)
                    menu_price = menu_name_to_price.get(product_key)
                    
                    # Nếu giá trong file khác với giá trong menu
                    # VÀ giá có biến số ở 3 chữ số cuối (không phải số tròn)
                    last_3_digits = price_int % 1000
                    
                    if last_3_digits != 0:  # Có biến số ở 3 chữ số cuối
                        # Tính ngược lại giá gốc: giá_thay_thế * 1.08 / 1.1 = giá_bia_gốc
                        estimated_original_price = round(price_float * 1.08 / 1.1)
                        
                        # Kiểm tra xem giá gốc có phải là số tròn (chia hết cho 1000) không
                        is_round_price = (estimated_original_price % 1000 == 0)
                        
                        # Nếu giá gốc là số tròn và trong khoảng hợp lý
                        if is_round_price and 20000 <= estimated_original_price <= 500000:
                            # Tìm món bia/rượu/Coke trong menu có giá gần với giá gốc tính được
                            # Tìm trong khoảng ±5000 để bao quát các giá khác nhau
                            original_beverage = None
                            best_match = None
                            best_diff = float('inf')
                            
                            for menu_price, items in menu_by_price.items():
                                # Xét các giá trong khoảng ±10000 để bao quát hơn
                                if abs(menu_price - estimated_original_price) <= 10000:
                                    for item in items:
                                        item_name = item.get('name', '')
                                        if is_beverage_with_10_percent_tax(item_name):
                                            # Tính giá thay thế từ giá menu này
                                            calc_replacement = round(menu_price * 1.10 / 1.08)
                                            diff = abs(price_float - calc_replacement)
                                            
                                            # Tìm món có giá thay thế khớp nhất với giá trong file
                                            # Ưu tiên khớp chính xác (sai số <= 2), nhưng cũng chấp nhận gần đúng (sai số <= 10000)
                                            if diff <= 10000 and diff < best_diff:
                                                best_diff = diff
                                                best_match = menu_price
                                                original_beverage = item
                            
                            # Nếu tìm thấy món bia/rượu/Coke với giá thay thế khớp (sai số <= 10000)
                            # Ưu tiên khớp chính xác (sai số <= 2), nhưng cũng chấp nhận gần đúng (sai số <= 10000)
                            if original_beverage is not None and best_diff <= 10000:
                                original_name = original_beverage.get('name', f'Bia/Rượu/Coke giá {best_match:,}đ')
                                original_beverage_price = best_match  # Sử dụng giá thực tế từ menu
                                
                                # Xác định loại: bia/rượu hay Coke (dựa vào giá và tên món)
                                original_name_lower = original_name.lower()
                                if 'coke' in original_name_lower or 'coca' in original_name_lower:
                                    item_type = "Coke (10% đường)"
                                elif 'sangria' in original_name_lower or 'wine' in original_name_lower or 'rượu' in original_name_lower:
                                    item_type = "Rượu"
                                else:
                                    item_type = "Bia/Rượu"
                                
                                beverages_replaced.append({
                                    'product': product_name_str,
                                    'price': price_float,
                                    'original_beverage_name': original_name,
                                    'original_beverage_price': original_beverage_price,
                                    'item_type': item_type
                                })
            
            # Lưu kết quả
            if has_unformatted:
                invoices_without_format.append(invoice_number)
            
            if beverages_replaced:
                invoices_with_beverages.append({
                    'invoice_number': invoice_number,
                    'file': invoice_file.name,
                    'replacements': beverages_replaced
                })
        
        except Exception as e:
            print(f"⚠️  Lỗi khi kiểm tra file {invoice_file.name}: {e}")
            continue
    
    # Trả về kết quả
    return {
        'invoices_without_format': invoices_without_format,
        'invoices_with_beverages': invoices_with_beverages
    }

def main():
    """Hàm chính"""
    print("="*60)
    print("🔍 KIỂM TRA HÓA ĐƠN (So sánh với menu gốc)")
    print("="*60)
    
    results = check_invoices()
    
    print(f"\n📋 Hóa đơn có món chưa format 'Tiếng Việt / Tiếng Anh':")
    print(f"   Tổng cộng: {len(results['invoices_without_format'])} hóa đơn")
    if results['invoices_without_format']:
        for inv_num in results['invoices_without_format']:
            print(f"   - {inv_num}")
    
    print(f"\n🍺 Hóa đơn có bia/rượu đã được thay thế:")
    print(f"   Tổng cộng: {len(results['invoices_with_beverages'])} hóa đơn")
    if results['invoices_with_beverages']:
        for item in results['invoices_with_beverages']:
            print(f"   - Hóa đơn {item['invoice_number']} ({item['file']}):")
            for replacement in item['replacements']:
                print(f"     • {replacement['product']} - {replacement['price']:,.0f}đ")
                print(f"       ← Thay thế từ: {replacement['original_beverage_name']} ({replacement['original_beverage_price']:,}đ)")
    
    print("="*60)
    
    return results

if __name__ == "__main__":
    main()
