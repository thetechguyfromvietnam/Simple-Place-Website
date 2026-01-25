#!/usr/bin/env python3
"""
Script kiểm tra thiếu món và sai số lượng giữa file input và output
"""

import re
from pathlib import Path
import openpyxl
import sys

def count_items_in_html_row(row, invoice_id):
    """Đếm số món trong 1 row HTML"""
    items = []
    cells = re.findall(r'<td[^>]*>(.*?)</td>', row)
    cells = [re.sub(r'<[^>]+>', '', cell).strip() for cell in cells]
    
    for i in range(len(cells) - 3):
        name = cells[i]
        qty_candidate = cells[i + 1]
        unit_candidate = cells[i + 2]
        price_candidate = cells[i + 3]
        
        # Kiểm tra điều kiện parse (giống logic trong process_invoices.py)
        if not (qty_candidate.isdigit() and 1 <= int(qty_candidate) <= 200):
            continue
        
        qty = int(qty_candidate)
        price_clean = price_candidate.replace(' ', '').replace(',', '').replace('.', '')
        if not price_clean.isdigit():
            continue
        
        price_value = float(price_clean)
        unit = unit_candidate if unit_candidate and not unit_candidate.isdigit() else 'Phần'
        
        if (len(name) < 2 or name.isdigit() or 
            name in ['', 'STT', 'Mã hoá đơn', 'Simple Place']):
            continue
        
        skip_patterns = [
            r'\bcrispy\b', r'\bsoft\b', r'cut in 4', r'- edit\s*$',
            r'đổi phương thức', r'\bpayment\b', r'\btransfer\b',
            r'\bcod\b', r'\batm\b', 'background-color', 'vertical-align',
            'ghi chú', 'giảm sốt'
        ]
        if any(re.search(pattern, name.lower()) for pattern in skip_patterns):
            continue
        
        if (price_value >= 500 and price_value <= 2000000 and 
            qty >= 1 and qty <= 200 and len(name) > 2):
            items.append({
                'name': name,
                'quantity': qty,
                'price': price_value,
                'unit': unit
            })
    
    return items

def count_items_in_html(content, invoice_id):
    """Đếm số món trong HTML cho 1 hóa đơn"""
    rows = content.split('<tr>')
    invoice_start_idx = None
    
    # Tìm row có số hóa đơn
    for i, row in enumerate(rows):
        if re.search(r'rowspan="\d+">' + invoice_id + r'</td>', row):
            invoice_start_idx = i
            break
    
    if invoice_start_idx is None:
        return []
    
    all_items = []
    parsed_positions = set()  # Track vị trí đã parse trong mỗi row
    
    # Parse các row thuộc hóa đơn này
    for i in range(invoice_start_idx, len(rows)):
        row = rows[i]
        
        # Kiểm tra xem có phải hóa đơn mới không
        if i > invoice_start_idx:
            if re.search(r'rowspan="\d+">(\d{6})</td>', row):
                break
        
        # Parse items trong row này
        cells = re.findall(r'<td[^>]*>(.*?)</td>', row)
        cells = [re.sub(r'<[^>]+>', '', cell).strip() for cell in cells]
        
        parsed_in_row = set()
        
        for j in range(len(cells) - 3):
            name = cells[j]
            qty_candidate = cells[j + 1]
            unit_candidate = cells[j + 2]
            price_candidate = cells[j + 3]
            
            # Kiểm tra điều kiện parse
            if not (qty_candidate.isdigit() and 1 <= int(qty_candidate) <= 200):
                continue
            
            qty = int(qty_candidate)
            price_clean = price_candidate.replace(' ', '').replace(',', '').replace('.', '')
            if not price_clean.isdigit():
                continue
            
            price_value = float(price_clean)
            unit = unit_candidate if unit_candidate and not unit_candidate.isdigit() else 'Phần'
            
            if (len(name) < 2 or name.isdigit() or 
                name in ['', 'STT', 'Mã hoá đơn', 'Simple Place']):
                continue
            
            skip_patterns = [
                r'\bcrispy\b', r'\bsoft\b', r'cut in 4', r'- edit\s*$',
                r'đổi phương thức', r'\bpayment\b', r'\btransfer\b',
                r'\bcod\b', r'\batm\b', 'background-color', 'vertical-align',
                'ghi chú', 'giảm sốt'
            ]
            if any(re.search(pattern, name.lower()) for pattern in skip_patterns):
                continue
            
            if (price_value >= 500 and price_value <= 2000000 and 
                qty >= 1 and qty <= 200 and len(name) > 2):
                
                # Check duplicate dựa trên vị trí cell (giống logic mới)
                if j in parsed_in_row:
                    continue
                parsed_in_row.add(j)
                
                all_items.append({
                    'name': name,
                    'quantity': qty,
                    'price': price_value,
                    'unit': unit
                })
    
    return all_items

def count_items_in_excel(excel_file):
    """Đếm số món trong file Excel"""
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        items = []
        
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=3).value  # Ten_san_pham
            quantity = ws.cell(row=row, column=5).value  # So_luong
            price = ws.cell(row=row, column=6).value  # Don_gia
            
            if not name:
                continue
            
            # Bỏ qua header
            if str(name).strip().lower() in ['ten_san_pham', 'tên sản phẩm', 'ten san pham']:
                continue
            
            try:
                qty = float(quantity) if quantity else 0
                prc = float(price) if price else 0
                
                if qty > 0 and prc > 0:
                    items.append({
                        'name': str(name).strip(),
                        'quantity': int(qty),
                        'price': prc
                    })
            except (ValueError, TypeError):
                continue
        
        return items
    except Exception as e:
        print(f"   ❌ Lỗi đọc Excel: {e}")
        return []

def normalize_name_for_comparison(name):
    """Normalize tên món để so sánh (lấy phần tiếng Anh, bỏ qua format)"""
    if not name:
        return ""
    
    # Lấy phần tiếng Anh nếu có format "Vietnamese / English"
    if ' / ' in name:
        name = name.split(' / ')[-1].strip()
    
    # Loại bỏ các ký tự đặc biệt và chuyển về lowercase
    name = re.sub(r'[^\w\s]', '', name.lower().strip())
    
    # Loại bỏ các từ thừa
    name = re.sub(r'\s*\(spicy\)\s*', '', name)
    name = re.sub(r'\s*\(.*?\)\s*', '', name)  # Bỏ tất cả text trong ngoặc
    name = re.sub(r'\s+extra\s*$', '', name)
    name = re.sub(r'\s+tacos?\s*$', '', name)  # Bỏ "tacos" ở cuối
    name = re.sub(r'\s+taco\s*$', '', name)
    
    return name.strip()

def match_items_by_price_and_name(input_items, output_items):
    """Match items dựa trên giá và tên (normalized)"""
    matched = []
    unmatched_input = []
    unmatched_output = list(output_items)
    
    for input_item in input_items:
        input_price = input_item['price']
        input_name_norm = normalize_name_for_comparison(input_item['name'])
        input_qty = input_item['quantity']
        
        best_match = None
        best_score = 0
        
        for output_item in unmatched_output:
            output_price = output_item['price']
            output_name_norm = normalize_name_for_comparison(output_item['name'])
            output_qty = output_item['quantity']
            
            # So sánh giá (cho phép sai số nhỏ do làm tròn)
            price_diff = abs(input_price - output_price)
            if price_diff > 100:  # Sai số > 100đ là khác món
                continue
            
            # So sánh tên (normalized)
            if input_name_norm and output_name_norm:
                # Tính điểm match
                input_words = set(input_name_norm.split())
                output_words = set(output_name_norm.split())
                
                if input_words and output_words:
                    common_words = input_words & output_words
                    score = len(common_words) / max(len(input_words), len(output_words))
                    
                    # Bonus nếu tên gần giống nhau
                    if input_name_norm in output_name_norm or output_name_norm in input_name_norm:
                        score += 0.3
                    
                    if score > best_score and score >= 0.3:  # Threshold 30%
                        best_score = score
                        best_match = output_item
            elif price_diff < 1:  # Nếu không match được tên nhưng giá giống hệt
                best_score = 1.0
                best_match = output_item
        
        if best_match:
            matched.append({
                'input': input_item,
                'output': best_match,
                'input_qty': input_qty,
                'output_qty': best_match['quantity'],
                'price': input_price
            })
            unmatched_output.remove(best_match)
        else:
            unmatched_input.append(input_item)
    
    return matched, unmatched_input, unmatched_output

def compare_items(input_items, output_items):
    """So sánh items giữa input và output"""
    issues = []
    
    # Match items
    matched, unmatched_input, unmatched_output = match_items_by_price_and_name(input_items, output_items)
    
    # Kiểm tra số lượng khác nhau trong các món đã match
    for match in matched:
        if match['input_qty'] != match['output_qty']:
            issues.append({
                'name': match['input']['name'],
                'matched_name': match['output']['name'],
                'price': match['price'],
                'input_qty': match['input_qty'],
                'output_qty': match['output_qty'],
                'diff': match['input_qty'] - match['output_qty'],
                'type': 'quantity_mismatch'
            })
    
    # Các món trong input nhưng không có trong output (thiếu)
    for item in unmatched_input:
        issues.append({
            'name': item['name'],
            'matched_name': None,
            'price': item['price'],
            'input_qty': item['quantity'],
            'output_qty': 0,
            'diff': item['quantity'],
            'type': 'missing'
        })
    
    # Các món trong output nhưng không có trong input (thừa - có thể là món thay thế bia/rượu)
    for item in unmatched_output:
        # Bỏ qua nếu giá là giá đã được điều chỉnh (có thể là món thay thế bia/rượu)
        # Giá thay thế thường có dạng: round(original_price * 1.10 / 1.08)
        # Kiểm tra xem có phải giá thay thế không
        is_replacement = False
        for input_item in input_items:
            # Nếu giá output gần với giá input * 1.10 / 1.08 (làm tròn)
            estimated_replacement = round(input_item['price'] * 1.10 / 1.08)
            if abs(item['price'] - estimated_replacement) <= 2:
                is_replacement = True
                break
        
        if not is_replacement:
            issues.append({
                'name': None,
                'matched_name': item['name'],
                'price': item['price'],
                'input_qty': 0,
                'output_qty': item['quantity'],
                'diff': -item['quantity'],
                'type': 'extra'
            })
    
    return issues

def check_all_invoices():
    """Kiểm tra tất cả hóa đơn"""
    base_dir = Path(__file__).parent
    data_dir = base_dir / "data"
    tax_dir = base_dir / "tax_files"
    
    # Tìm các file input
    input_files = []
    if data_dir.exists():
        input_files = list(data_dir.glob("*.xls")) + list(data_dir.glob("*.html"))
    
    if not input_files:
        print("❌ Không tìm thấy file input trong thư mục data/")
        return
    
    print("=" * 80)
    print("🔍 KIỂM TRA THIẾU MÓN VÀ SAI SỐ LƯỢNG")
    print("=" * 80)
    print()
    
    # Đọc tất cả file input
    all_html_content = ""
    for input_file in input_files:
        print(f"📂 Đang đọc file: {input_file.name}")
        try:
            with open(input_file, 'r', encoding='utf-8', errors='ignore') as f:
                all_html_content += f.read() + "\n"
        except Exception as e:
            print(f"   ❌ Lỗi: {e}")
    
    # Tìm tất cả số hóa đơn trong HTML
    rows = all_html_content.split('<tr>')
    invoice_ids = set()
    for row in rows:
        match = re.search(r'rowspan="\d+">(\d{6})</td>', row)
        if match:
            invoice_ids.add(match.group(1))
    
    print(f"📊 Tìm thấy {len(invoice_ids)} hóa đơn trong file input")
    print()
    
    # Kiểm tra từng hóa đơn
    issues_found = []
    invoices_checked = 0
    
    for invoice_id in sorted(invoice_ids):
        # Tìm file Excel tương ứng
        matching_files = list(tax_dir.glob(f"{invoice_id} - *"))
        if not matching_files:
            continue
        
        excel_file = matching_files[0]
        invoices_checked += 1
        
        # Đếm món trong HTML
        input_items = count_items_in_html(all_html_content, invoice_id)
        
        # Đếm món trong Excel
        output_items = count_items_in_excel(excel_file)
        
        # So sánh
        issues = compare_items(input_items, output_items)
        
        if issues:
            issues_found.append({
                'invoice_id': invoice_id,
                'input_count': len(input_items),
                'output_count': len(output_items),
                'input_total_qty': sum(item['quantity'] for item in input_items),
                'output_total_qty': sum(item['quantity'] for item in output_items),
                'issues': issues
            })
    
    print("=" * 80)
    print("📊 KẾT QUẢ KIỂM TRA")
    print("=" * 80)
    print()
    print(f"✅ Đã kiểm tra {invoices_checked} hóa đơn")
    print()
    
    if issues_found:
        print(f"⚠️  Tìm thấy {len(issues_found)} hóa đơn có vấn đề:")
        print()
        
        for issue in issues_found:
            print(f"❌ Hóa đơn {issue['invoice_id']}:")
            print(f"   Input: {issue['input_count']} món, tổng số lượng: {issue['input_total_qty']}")
            print(f"   Output: {issue['output_count']} món, tổng số lượng: {issue['output_total_qty']}")
            print(f"   Số món bị sai: {len(issue['issues'])}")
            print()
            
            for item_issue in issue['issues'][:5]:  # Chỉ hiển thị 5 món đầu
                if item_issue['diff'] > 0:
                    print(f"      ⚠️  Thiếu: {item_issue['name']} (Giá: {item_issue['price']:,.0f}đ)")
                    print(f"         Input: {item_issue['input_qty']} | Output: {item_issue['output_qty']} | Thiếu: {item_issue['diff']}")
                else:
                    print(f"      ⚠️  Thừa: {item_issue['name']} (Giá: {item_issue['price']:,.0f}đ)")
                    print(f"         Input: {item_issue['input_qty']} | Output: {item_issue['output_qty']} | Thừa: {abs(item_issue['diff'])}")
            
            if len(issue['issues']) > 5:
                print(f"      ... và {len(issue['issues']) - 5} món khác")
            print()
    else:
        print("✅ TẤT CẢ HÓA ĐƠN ĐỀU ĐÚNG!")
        print("   Không có món nào bị thiếu hoặc sai số lượng")
    
    print("=" * 80)

if __name__ == '__main__':
    check_all_invoices()
