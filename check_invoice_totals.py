#!/usr/bin/env python3
"""
Kiểm tra tổng tiền các món trong tất cả file hóa đơn
Xem có khớp với tổng tiền trong tên file không
"""

import openpyxl
from pathlib import Path
import re

def extract_total_from_filename(filename):
    """Trích xuất tổng tiền từ tên file (VD: 240002 - transfer - 642.600đ.xlsx -> 642600)"""
    # Tìm số cuối cùng trước "đ" hoặc "đ.xlsx"
    match = re.search(r'([\d.,]+)\s*đ', filename)
    if match:
        total_str = match.group(1).replace('.', '').replace(',', '')
        try:
            return float(total_str)
        except:
            return None
    return None

def check_invoice_totals():
    """Kiểm tra tổng tiền tất cả file hóa đơn"""
    base_dir = Path(__file__).parent
    tax_dir = base_dir / "tax_files"
    
    if not tax_dir.exists():
        print("❌ Thư mục tax_files không tồn tại!")
        return
    
    print("=" * 80)
    print("🔍 KIỂM TRA TỔNG TIỀN CÁC MÓN TRONG HÓA ĐƠN")
    print("=" * 80)
    print()
    
    issues = []
    total_files = 0
    correct_files = 0
    
    for invoice_file in sorted(tax_dir.glob("*.xlsx")):
        if invoice_file.name.startswith('.'):
            continue
            
        total_files += 1
        try:
            wb = openpyxl.load_workbook(invoice_file)
            ws = wb.active
            
            # Lấy số hóa đơn từ tên file
            invoice_number = invoice_file.stem.split(' - ')[0] if ' - ' in invoice_file.stem else invoice_file.stem
            
            # Trích xuất tổng tiền từ tên file (đã bao gồm VAT 8%)
            expected_total_with_vat = extract_total_from_filename(invoice_file.name)
            
            # Tính tổng tiền các món từ file Excel
            items_total = 0.0
            items_count = 0
            
            for row in range(2, ws.max_row + 1):
                product_name = ws.cell(row=row, column=3).value  # Ten_san_pham
                quantity = ws.cell(row=row, column=5).value      # So_luong
                price = ws.cell(row=row, column=6).value         # Don_gia
                
                if not product_name:
                    continue
                
                # Bỏ qua header nếu có
                if str(product_name).strip().lower() in ['ten_san_pham', 'tên sản phẩm', 'ten san pham']:
                    continue
                
                try:
                    qty = float(quantity) if quantity else 0
                    prc = float(price) if price else 0
                    item_total = qty * prc
                    items_total += item_total
                    items_count += 1
                except (ValueError, TypeError):
                    continue
            
            # Tính tổng tiền có VAT (items_total * 1.08)
            calculated_total_with_vat = items_total * 1.08
            
            # So sánh với tổng tiền trong tên file
            if expected_total_with_vat:
                diff = abs(calculated_total_with_vat - expected_total_with_vat)
                diff_percent = (diff / expected_total_with_vat * 100) if expected_total_with_vat > 0 else 0
                
                # Cho phép sai số nhỏ (do làm tròn)
                tolerance = 1.0  # 1 VND
                
                if diff > tolerance:
                    issues.append({
                        'file': invoice_file.name,
                        'invoice_number': invoice_number,
                        'expected': expected_total_with_vat,
                        'calculated': calculated_total_with_vat,
                        'items_total': items_total,
                        'diff': diff,
                        'diff_percent': diff_percent,
                        'items_count': items_count
                    })
                else:
                    correct_files += 1
            else:
                # Không tìm thấy tổng tiền trong tên file
                issues.append({
                    'file': invoice_file.name,
                    'invoice_number': invoice_number,
                    'expected': None,
                    'calculated': calculated_total_with_vat,
                    'items_total': items_total,
                    'diff': None,
                    'diff_percent': None,
                    'items_count': items_count,
                    'error': 'Không tìm thấy tổng tiền trong tên file'
                })
        
        except Exception as e:
            issues.append({
                'file': invoice_file.name,
                'invoice_number': invoice_file.stem,
                'error': f'Lỗi khi đọc file: {str(e)}'
            })
    
    # In kết quả
    print(f"📊 Tổng số file kiểm tra: {total_files}")
    print(f"✅ File đúng: {correct_files}")
    print(f"❌ File có vấn đề: {len(issues)}")
    print()
    
    # Hiển thị thống kê chi tiết
    if total_files > 0:
        print("=" * 80)
        print("📈 THỐNG KÊ CHI TIẾT:")
        print("=" * 80)
        
        # Tính lại để lấy thông tin chi tiết
        all_totals = []
        all_items_counts = []
        
        for invoice_file in sorted(tax_dir.glob("*.xlsx")):
            if invoice_file.name.startswith('.'):
                continue
            try:
                wb = openpyxl.load_workbook(invoice_file)
                ws = wb.active
                
                items_total = 0.0
                items_count = 0
                
                for row in range(2, ws.max_row + 1):
                    product_name = ws.cell(row=row, column=3).value
                    quantity = ws.cell(row=row, column=5).value
                    price = ws.cell(row=row, column=6).value
                    
                    if not product_name:
                        continue
                    
                    if str(product_name).strip().lower() in ['ten_san_pham', 'tên sản phẩm', 'ten san pham']:
                        continue
                    
                    try:
                        qty = float(quantity) if quantity else 0
                        prc = float(price) if price else 0
                        item_total = qty * prc
                        items_total += item_total
                        items_count += 1
                    except (ValueError, TypeError):
                        continue
                
                if items_total > 0:
                    all_totals.append(items_total * 1.08)
                    all_items_counts.append(items_count)
            except:
                pass
        
        if all_totals:
            print(f"   💰 Tổng tiền nhỏ nhất (có VAT): {min(all_totals):,.0f} VND")
            print(f"   💰 Tổng tiền lớn nhất (có VAT): {max(all_totals):,.0f} VND")
            print(f"   💰 Tổng tiền trung bình (có VAT): {sum(all_totals)/len(all_totals):,.0f} VND")
            print(f"   📦 Số món trung bình: {sum(all_items_counts)/len(all_items_counts):.1f}")
            print(f"   📦 Số món ít nhất: {min(all_items_counts)}")
            print(f"   📦 Số món nhiều nhất: {max(all_items_counts)}")
        print()
    
    if issues:
        print("=" * 80)
        print("❌ CÁC FILE CÓ VẤN ĐỀ:")
        print("=" * 80)
        print()
        
        for issue in issues:
            print(f"📄 File: {issue['file']}")
            print(f"   Mã HĐ: {issue['invoice_number']}")
            
            if 'error' in issue:
                print(f"   ❌ {issue['error']}")
            else:
                if issue['expected'] is None:
                    print(f"   ⚠️  Không tìm thấy tổng tiền trong tên file")
                    print(f"   💰 Tổng tính được (có VAT): {issue['calculated']:,.2f} VND")
                else:
                    print(f"   💰 Tổng trong tên file:     {issue['expected']:,.2f} VND")
                    print(f"   💰 Tổng tính được (có VAT): {issue['calculated']:,.2f} VND")
                    print(f"   📊 Tổng các món (chưa VAT): {issue['items_total']:,.2f} VND")
                    print(f"   ⚠️  Chênh lệch:              {issue['diff']:,.2f} VND ({issue['diff_percent']:.2f}%)")
                
                print(f"   📦 Số món: {issue.get('items_count', 0)}")
            
            print()
    else:
        print("=" * 80)
        print("✅ TẤT CẢ FILE ĐỀU ĐÚNG!")
        print("=" * 80)
    
    return issues

if __name__ == '__main__':
    check_invoice_totals()
